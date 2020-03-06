from install_and_import import install_and_import
from abc import ABC, abstractmethod
install_and_import('openpyxl')
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import calibre
import upc
import logging
import my_logger
import re
install_and_import('win32com.client')
install_and_import('dropbox')
import win32com.client
import dropbox

win32c = win32com.client.constants


class MyWorkbook:
    """A workbook class that wraps openpyxl wb"""

    # Loads the workbook, if there isn't a file name provided, load an empty workbook
    def open_workbook(self):
        try:
            return load_workbook(filename=self.file_path + self.file)
        except Exception as inst:
            # If we can't open the spreadsheet, creating a new spreadsheet
            print("Unable to open file", self.file, "error:", inst, "making new file")
            self.logger.warning("Can't open file %s error: %s, making new file", self.file, inst)
            temp = Workbook()
            temp.save(filename=self.file_path + self.file)
            return temp

    def __init__(self, file, file_path):
        self.logger = logging.getLogger("Workbook")
        self.file = file
        self.file_path = file_path
        self.ws = None
        self.pyxlwb = self.open_workbook()

    def open_worksheet(self, name, sort_fields, book=False, movie=False):
        if book:
            self.ws = MyBookWorksheet(self, name, sort_fields)
        elif movie:
            self.ws = MyMovieWorksheet(self, name, sort_fields)
        return self.ws

    # Uploads the spreadsheet to dropbox
    def upload_to_dropbox(self, token):
        #  target location in Dropbox
        target = "/"
        #  the target folder
        filename = self.file
        targetfile = target + filename  # the target path and file name
        self.logger.debug("Attempting to upload %s to %s", self.file, targetfile)
        #  Create a dropbox object using an API v2 key
        d = dropbox.Dropbox(token)
        #  open the file and upload it
        with open(self.file_path + self.file, 'rb') as f:
            #  upload gives you metadata about the file
            #  we want to overwrite any previous version of the file
            d.files_upload(f.read(), targetfile, mode=dropbox.files.WriteMode("overwrite"))
            self.logger.info("Uploaded the file")
        #  create a shared link
        try:
            link = d.sharing_create_shared_link_with_settings(targetfile)
            url = link.url
        except dropbox.exceptions.ApiError as apierror:
            self.logger.info("API Error %s", apierror)
            self.logger.info("Shared Link already exists, getting shared link")
            try:
                link = d.sharing_list_shared_links(targetfile, direct_only=True)
                url = link.links[0].url
            except dropbox.exceptions.ApiError as apierror:
                self.logger.warning("Unable to get share link, upload may have failed error was: %s", apierror)
                return

        #  link which directly downloads by replacing ?dl=0 with ?dl=1
        dl_url = re.sub(r"\?dl\=0", "?dl=1", url)
        self.logger.info("File can be accessed at %s", dl_url)
        print(dl_url)

    # Saves the Spreadsheet, attempts to sort and resize if possible, as well as make the worksheet a table
    def save_and_close(self, token):
        try:
            self.ws.get_table()
            self.ws.resize_columns()
            self.save()
        except Exception as inst:
            # If the Table, resizing or saving fails, report the error
            # and give one more chance for just the save to succeed
            self.logger.exception("Can't save file %s error: %s", self.file_path + self.file, inst)
            print("Can't save file", self.file_path + self.file, "Error:", inst)
            print("nRemoving resizing features and sorting to increase change of success\a")
            input("You have one more chance, fix error and press enter to save again")
            self.save()
            print("Success!")
            self.logger.critical("Save succeeded")
        try:
            # Close openpyxl's connection to the spreadsheet and attempt to sort it
            self.pyxlwb.close()
            self.ws.sort()
            self.logger.debug("Excel file sorted")
        except Exception as inst:
            # If this fails, just move on
            # it may fail due to excel not being installed, odd format, or a lot of other issues, it's a fragile process
            self.logger.exception("Can't sort file %s error: %s", self.file_path + self.file, inst)
            print("The file was not sorted\a")
        if token:
            # If there is a dropbox token, upload the file to dropbox
            self.logger.info("Trying to upload to dropbox")
            self.upload_to_dropbox(token)
            print("File was uploaded to dropbox")
        else:
            self.logger.warning("File failed to upload, possibly due to no token")
            print("File was not uploaded to dropbox")

    # Save the workbook locally
    def save(self):
        try:
            self.pyxlwb.save(filename=self.file_path + self.file)
            self.logger.info("Excel file saved")
        # If the save fails, report the error and give one more chance for success
        except Exception as inst:
            self.logger.exception("Can't save file %s error: %s", self.file_path + self.file, inst)
            print("Can't save file", self.file_path + self.file, "Error:", inst)
            input("You have one more chance, fix error and press enter to save again\a")
            self.pyxlwb.save(filename=self.file_path + self.file)
            print("Success!")
            self.logger.critical("Save succeeded")


class MyWorksheet(ABC):
    """A Class that wraps openpyxl worksheet class"""

    @abstractmethod
    def __init__(self, wb, name, sort_fields):
        self.logger = logging.getLogger("Worksheet")
        self.wb = wb
        self.name = name
        self.sort_fields = sort_fields
        self.table = None
        self.dims = {}
        self.pyxlws = None
        super().__init__()

    # Convert the worksheet to a table, if a table is there ensure the dimensions are correct
    def get_table(self):
        self.logger.debug("Checking Table")
        if self.table and self.table.name == self.name:
            self.logger.info("Table exists, returning reference")
            return self.table
        self.logger.debug("Checking ws for table")
        tables = self.pyxlws._tables
        # Check all tables to see if the one we want exists
        for table in tables:
            if table.name == "My"+self.name:
                self.table = table
        if self.table is None:
            if len(self.pyxlws['A']) > 1:
                # If it doesn't, create it
                self.logger.info("Table does not exist, making")
                style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                self.table = Table(displayName="My"+self.name, ref=self.pyxlws.dimensions, tableStyleInfo=style)
                self.pyxlws.add_table(self.table)
            else:
                self.logger.warning("Table is to small to be created, skipping")
        else:
            # if it does, set it's dimensions to the dimensions of the work sheet
            self.logger.info("Table already exists, updating dimensions")
            self.table.ref = self.pyxlws.dimensions
        return self.table

    # Sorting can't be done by openpyxl
    # so this will attempt to open it in excel and sort from there, assumes the file is a table
    def sort(self):
        excel = None
        if self.get_table() is not None:
            try:
                self.logger.info("Attempting to sort columns")
                # Open Excel
                self.logger.debug("opening excel")
                excel = win32com.client.Dispatch("Excel.Application")
                filepath = self.wb.file_path + self.wb.file
                filepath = filepath.replace('\\', '\\\\')
                self.logger.debug("Attempting to open %s", filepath)
                # Load file and Worksheet
                excel_wb = excel.Workbooks.Open(Filename=filepath)
                self.logger.debug("using worksheet %s", self.name)
                excel_ws = excel_wb.Worksheets(self.name)
                # Grab Table
                tab = excel_ws.ListObjects(self.table.name)
                # Remove current sorts, and sort by sort_fields
                tab.Sort.SortFields.Clear()
                for field in self.sort_fields:
                    tab.Sort.SortFields.Add(Key=excel_ws.Range(self.table.name+'[' + field + ']'), Order=1)
                # Apply Sort, Save, Then close Excel
                tab.Sort.Apply()
                self.logger.debug("Finished Sorting, Saving Changes")
                excel_wb.Save()
                self.logger.debug("Exiting Excel")
                excel.Application.Quit()
                return self
            except Exception as inst:
                # If this fails for any reason close without saving and move on
                self.logger.exception("Can't sort file %s error: %s", self.wb.file, inst)
                if excel:
                    excel.Application.Quit()
                raise inst
        self.logger.warning("Skipping Sorting as table does not exist")

    # Auto adjust columns to be wide enough for the widest text
    def resize_columns(self):
        self.dims = {}
        self.logger.debug("Attempting to resize the columns to fit text length")
        # Check each cell for its width
        for row in self.pyxlws.rows:
            for cell in row:
                if cell.value:
                    # if it's width is larger than the current dimension, increase the dimension
                    self.dims[cell.column_letter] = max((self.dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in self.dims.items():
            # Apply maximums to the columns
            self.pyxlws.column_dimensions[col].width = value
        return self

    # Takes the worksheet and metadata and adds it to the worksheet
    def add_to_db(self, metadata):
        self.logger.info("Adding %s to worksheet", metadata)
        self.logger.debug("Adding row to end of worksheet")
        self.pyxlws.append(metadata)
        return metadata

    # Checks if the barcode is already in the worksheet
    def check_duplicate(self, barcode):
        self.logger.debug("Checking that %s is not already in workbook", barcode)
        for i, row in enumerate(self.pyxlws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True)):
            for cell in row:
                if cell == barcode:
                    self.logger.warning("Barcode %s found in row %s, skipping", barcode, i)
                    return True
        self.logger.debug("Barcode %s not found, adding entry", barcode)
        return False

    # Main function that kicks off checking for duplicate, gathering metadata, and adding it to spreadsheet
    @abstractmethod
    def add_entry(self, client, barcode):
        thread_log_handler = my_logger.start_thread_logging(self.wb.file_path)
        return thread_log_handler

    @abstractmethod
    def incomplete(self, result, previous):
        pass


class MyBookWorksheet(MyWorksheet):
    """Worksheet devoted to tracking books"""

    def __init__(self, wb, name, sort_fields):
        super().__init__(wb, name, sort_fields)
        try:
            self.pyxlws = self.wb.pyxlwb[name]
        except Exception as inst:
            print("Unable to open worksheet", name, "error:", inst, "making new worksheet")
            self.logger.warning("Can't open worksheet %s error: %s, making new worksheet", name, inst)
            if self.wb.pyxlwb.active.title == "Sheet":
                self.logger.debug("using existing sheet")
                self.pyxlws = self.wb.pyxlwb.active
                self.pyxlws.title = self.name
            else:
                self.logger.debug("Creating fresh sheet")
                self.pyxlws = self.wb.pyxlwb.create_sheet(self.name)
            # Add header row
            row = ("Author", "Title", "Series", "ISBN", "Location", "Condition", "Borrowed By", "Date Borrowed")
            self.pyxlws.append(row)
            self.wb.save()

    def add_entry(self, client, barcode):
        thread_log_handler = super().add_entry(client, barcode)
        search_logger = logging.getLogger("Worksheet")
        if not self.check_duplicate(barcode):
            metadata = calibre.get_metadata(client, barcode)
            search_logger.debug("Got metadata, adding to DB")
            my_logger.stop_thread_logging(thread_log_handler)
            return self.add_to_db(metadata)
        else:
            search_logger.info("%s is already in spreadsheet", barcode)
            my_logger.stop_thread_logging(thread_log_handler)
            return ["duplicate", "duplicate", "duplicate", "duplicate"]

    def incomplete(self, result, previous):
        blank = False
        blanks = None
        self.logger.debug("blanks are checked against %s", result)
        for line in result:
            # If any blank lines then open the blanks file
            if line[0] is None or line[1] is None:
                blank = True
        if blank:
            try:
                # Open missing.txt in append mode, create it if it doesn't exist
                blanks = open(self.wb.file_path + "missing.txt", "a+")
                for i, line in enumerate(result):
                    if line[0] is None or line[1] is None:
                        # Depending on information, write the file with the relevant info
                        self.logger.warning("Adding note about missing information on %s", line[3])
                        self.logger.debug("Full information is %s", line)
                        if i == 0:
                            if previous is None:
                                blanks.write("First item scanned, " + line[3] + " had missing data\n")
                            else:
                                blanks.write(
                                    line[3] + " which was scanned after " + previous[3] + ", written by " + previous[
                                        0] + " had missing data\n")
                        else:
                            blanks.write(
                                line[3] + " which was scanned after " + result[i - 1][3] + ", written by " +
                                result[i - 1][
                                    0] + " had missing data\n")
            finally:
                if blanks:
                    blanks.close()
                    return True
        else:
            self.logger.info("No blanks found in this set of results")
            return False


class MyMovieWorksheet(MyWorksheet):
    """Worksheet devoted to tracking movies"""

    def __init__(self, wb, name, sort_fields):
        super().__init__(wb, name, sort_fields)
        try:
            self.pyxlws = self.wb.pyxlwb[name]
        except Exception as inst:
            print("Unable to open worksheet", name, "error:", inst, "making new worksheet")
            self.logger.warning("Can't open worksheet %s error: %s, making new worksheet", name, inst)
            if self.wb.pyxlwb.active.title == "Sheet":
                self.logger.debug("using existing sheet")
                self.pyxlws = self.wb.pyxlwb.active
                self.pyxlws.title = self.name
            else:
                self.logger.debug("Creating fresh sheet")
                self.pyxlws = self.wb.pyxlwb.create_sheet(self.name)
            # Add header row
            row = ("Title", "Series", "Format", "UPC", "Borrowed By", "Date Borrowed")
            self.pyxlws.append(row)
            self.wb.save()

    def add_entry(self, client, barcode):
        thread_log_handler = super().add_entry(client, barcode)
        if not self.check_duplicate(barcode):
            metadata = upc.get_metadata(barcode)
            logging.debug("Got metadata, adding to DB")
            my_logger.stop_thread_logging(thread_log_handler)
            return self.add_to_db(metadata)
        else:
            logging.info("%s is already in spreadsheet", barcode)
            my_logger.stop_thread_logging(thread_log_handler)
            return ["duplicate", "duplicate", "duplicate", "duplicate"]

    def incomplete(self, result, previous):
        blank = False
        blanks = None
        for line in result:
            # If any blank lines then open the blanks file
            if line[0] is None or line[2] is None:
                blank = True
        if blank:
            try:
                # Open missing.txt in append mode, create it if it doesn't exist
                blanks = open(self.wb.file_path + "missing.txt", "a+")
                for i, line in enumerate(result):
                    if line[0] is None or line[2] is None:
                        # Depending on information, write the file with the relevant info
                        self.logger.warning("Adding note about missing information on %s", line[3])
                        self.logger.debug("Full information is %s", line)
                        if i == 0:
                            if previous is None:
                                blanks.write("First item scanned, " + line[3] + " had missing data\n")
                            else:
                                blanks.write(
                                    line[3] + " which was scanned after " + previous[3] + ", named " + previous[0] +
                                    " had missing data\n")
                        else:
                            blanks.write(line[3] + " which was scanned after " + result[i - 1][3] + ", named " +
                                         result[i - 1][0] + " had missing data\n")
            finally:
                if blanks:
                    blanks.close()
                    return True
        else:
            self.logger.info("No blanks found in this set of results")
            return False


# Downloads the spreadsheet from dropbox
def download_from_dropbox(file, token):
    # Updates the \\ to work properly
    # Extract filename from full path
    filename = file[file.rfind('\\') + 1:]
    location = file[:file.rfind('\\')]
    logger = logging.getLogger("Download_from_Dropbox")
    logger.debug("Attempting to download %s to %s", filename, location)
    try:
        # Attempt to download dropbox file
        d = dropbox.Dropbox(token)
        d.files_download_to_file(file, '/' + filename)
        logger.info("downloaded %s to %s", filename, location)
    except Exception as inst:
        # If it fails return a failure and move on
        logger.warning("Unable to download file from dropbox, error %s", inst)
        print("Unable to download file from dropbox, error:", inst, "\a")
        return 1
    return 0
